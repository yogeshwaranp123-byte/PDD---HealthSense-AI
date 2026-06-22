import os
import sys
import time
import asyncio
import httpx
import statistics
import subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Configuration
API_URL = "http://127.0.0.1:8000"
CONCURRENCY = int(os.environ.get("LOAD_TEST_CONCURRENCY", 100))
DURATION_SECONDS = int(os.environ.get("LOAD_TEST_DURATION", 60))
TEST_EMAIL = "tester@healthsense.ai"
TEST_PASSWORD = "Demo@HealthSense2026!"

# Paths relative to the project root
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND_DIR = os.path.join(PROJECT_ROOT, "backend")
UVICORN_LOG = os.path.join(PROJECT_ROOT, "tests", "uvicorn_load_test.log")

class RequestRecord:
    def __init__(self, endpoint, timestamp, latency_ms, status_code, success):
        self.endpoint = endpoint
        self.timestamp = timestamp  # time.perf_counter() relative to start
        self.latency_ms = latency_ms
        self.status_code = status_code
        self.success = success

def check_port_open(port):
    """Check if the local port is already open."""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def kill_port_owner(port):
    """Attempt to kill any process currently listening on the specified port on Windows."""
    try:
        output = subprocess.check_output(f'netstat -aon | findstr LISTENING | findstr :{port}', shell=True).decode()
        for line in output.strip().split('\n'):
            parts = line.strip().split()
            if len(parts) >= 5 and parts[1].endswith(f':{port}'):
                pid = parts[-1]
                print(f"[Init] Found process {pid} using port {port}. Terminating it...")
                subprocess.run(f'taskkill /F /PID {pid}', shell=True, check=True)
                time.sleep(1.0)
    except subprocess.CalledProcessError:
        pass  # Port is not in use or netstat failed

def start_backend_server():
    """Start uvicorn in a background process using the virtual environment."""
    print("\n" + "="*50)
    print("[Server] Starting backend FastAPI server...")
    print("="*50)
    
    # Clean up port 8000 if needed
    if check_port_open(8000):
        kill_port_owner(8000)
    
    # Path to virtual env python
    venv_python = os.path.join(BACKEND_DIR, "venv", "Scripts", "python.exe")
    if not os.path.exists(venv_python):
        # fallback to standard system python if venv not structured this way
        venv_python = "python"
        print(f"[Server] Virtualenv python not found at {venv_python}. Using default python.")
    
    print(f"[Server] Using Python: {venv_python}")
    print(f"[Server] Writing logs to: {UVICORN_LOG}")
    
    # Start process
    log_file = open(UVICORN_LOG, "w", encoding="utf-8")
    
    # Environment variables
    env = os.environ.copy()
    # Ensure stdout/stderr are unbuffered so they write to file immediately
    env["PYTHONUNBUFFERED"] = "1"
    
    proc = subprocess.Popen(
        [venv_python, "-m", "uvicorn", "app.main:app", "--host", "127.0.0.1", "--port", "8000"],
        cwd=BACKEND_DIR,
        stdout=log_file,
        stderr=subprocess.STDOUT,
        env=env
    )
    
    return proc, log_file

async def wait_for_server(timeout=15.0):
    """Wait until the local FastAPI server becomes responsive."""
    print("[Server] Waiting for API server to become ready...")
    start_time = time.time()
    async with httpx.AsyncClient() as client:
        while time.time() - start_time < timeout:
            try:
                resp = await client.get(f"{API_URL}/", timeout=1.0)
                if resp.status_code == 200:
                    print(f"[Server] FastAPI server is up and responsive! (took {time.time() - start_time:.2f}s)")
                    return True
            except (httpx.ConnectError, httpx.ConnectTimeout):
                pass
            await asyncio.sleep(0.5)
    return False

async def authenticate_user():
    """Perform login once to retrieve JWT token for Scenario B."""
    print("[Auth] Authenticating load test user...")
    async with httpx.AsyncClient() as client:
        try:
            resp = await client.post(
                f"{API_URL}/auth/login",
                json={"email": TEST_EMAIL, "password": TEST_PASSWORD},
                timeout=5.0
            )
            if resp.status_code == 200:
                data = resp.json()
                token = data.get("access_token")
                print(f"[Auth] Logged in successfully. Token acquired.")
                return token
            else:
                print(f"[Auth] Login failed with status {resp.status_code}: {resp.text}")
                return None
        except Exception as e:
            print(f"[Auth] Exception during authentication: {e}")
            return None

async def worker(client: httpx.AsyncClient, endpoint_path: str, headers: dict, results: list, stop_event: asyncio.Event, start_time_ref: float):
    """Simulate a single virtual user sending requests continuously."""
    url = f"{API_URL}{endpoint_path}"
    while not stop_event.is_set():
        req_start = time.perf_counter()
        timestamp = req_start - start_time_ref
        try:
            resp = await client.get(url, headers=headers, timeout=5.0)
            latency = (time.perf_counter() - req_start) * 1000.0  # ms
            success = 200 <= resp.status_code < 300
            status_code = resp.status_code
        except httpx.TimeoutException:
            latency = 5000.0
            success = False
            status_code = 504
        except Exception:
            latency = 0.0
            success = False
            status_code = 500
        
        results.append(RequestRecord(endpoint_path, timestamp, latency, status_code, success))
        
        # Small sleep yield to prevent tight-loop blocking of event loop
        await asyncio.sleep(0.005)

async def run_scenario(scenario_name: str, endpoint_path: str, headers: dict, duration: int, concurrency: int):
    """Run a load test scenario with the specified concurrency and duration."""
    print(f"\n" + "-"*60)
    print(f"[Test] Starting {scenario_name} ({endpoint_path})")
    print(f"[Test] Concurrency: {concurrency} Users | Duration: {duration} Seconds")
    print("-"*60)
    
    results = []
    stop_event = asyncio.Event()
    
    # We use a custom client with TCP connection pooling and keep-alive settings matching high concurrency
    limits = httpx.Limits(max_keepalive_connections=concurrency, max_connections=concurrency * 2)
    async with httpx.AsyncClient(limits=limits) as client:
        start_time_ref = time.perf_counter()
        
        # Spawn workers
        tasks = []
        for _ in range(concurrency):
            tasks.append(asyncio.create_task(worker(client, endpoint_path, headers, results, stop_event, start_time_ref)))
        
        # Report progress dynamically
        for elapsed in range(1, duration + 1):
            await asyncio.sleep(1.0)
            completed_requests = len(results)
            success_requests = sum(1 for r in results if r.success)
            current_rps = completed_requests / elapsed
            print(f"      Time: {elapsed:2d}s | Req: {completed_requests:5d} | Success: {success_requests:5d} | RPS: {current_rps:.1f}", end="\r")
            
        print(f"\n[Test] Time's up! Stopping workers...")
        stop_event.set()
        
        # Wait for all workers to shut down
        await asyncio.gather(*tasks, return_exceptions=True)
        print(f"[Test] Completed {scenario_name} Run. Gathered {len(results)} request records.")
        
    return results

def compute_metrics(name, records, total_duration):
    """Compute performance metrics from raw request records."""
    if not records:
        return {
            "name": name, "total_requests": 0, "successful_requests": 0, "failed_requests": 0,
            "success_rate": 0.0, "rps": 0.0, "avg_latency": 0.0, "min_latency": 0.0,
            "max_latency": 0.0, "std_dev": 0.0, "p50": 0.0, "p90": 0.0, "p95": 0.0, "p99": 0.0
        }
        
    latencies = [r.latency_ms for r in records if r.success]
    total_reqs = len(records)
    success_reqs = sum(1 for r in records if r.success)
    failed_reqs = total_reqs - success_reqs
    success_rate = (success_reqs / total_reqs) * 100.0 if total_reqs > 0 else 0.0
    rps = total_reqs / total_duration
    
    if latencies:
        avg_latency = statistics.mean(latencies)
        min_latency = min(latencies)
        max_latency = max(latencies)
        std_dev = statistics.stdev(latencies) if len(latencies) > 1 else 0.0
        
        # Calculate percentiles
        sorted_latencies = sorted(latencies)
        n = len(sorted_latencies)
        def percentile(p):
            idx = int(p * (n - 1))
            return sorted_latencies[idx]
            
        p50 = percentile(0.50)
        p90 = percentile(0.90)
        p95 = percentile(0.95)
        p99 = percentile(0.99)
    else:
        avg_latency = min_latency = max_latency = std_dev = p50 = p90 = p95 = p99 = 0.0
        
    return {
        "name": name,
        "total_requests": total_reqs,
        "successful_requests": success_reqs,
        "failed_requests": failed_reqs,
        "success_rate": success_rate,
        "rps": rps,
        "avg_latency": avg_latency,
        "min_latency": min_latency,
        "max_latency": max_latency,
        "std_dev": std_dev,
        "p50": p50,
        "p90": p90,
        "p95": p95,
        "p99": p99
    }

def get_time_series_data(records, total_duration):
    """Aggregate requests into 1-second bins to create a time-series history."""
    bins = [[] for _ in range(total_duration)]
    for r in records:
        second_idx = int(r.timestamp)
        if 0 <= second_idx < total_duration:
            bins[second_idx].append(r)
            
    time_series = []
    for idx, bin_records in enumerate(bins):
        sec = idx + 1
        req_count = len(bin_records)
        successful = sum(1 for r in bin_records if r.success)
        avg_lat = statistics.mean([r.latency_ms for r in bin_records if r.success]) if successful > 0 else 0.0
        time_series.append((sec, req_count, avg_lat))
    return time_series

def generate_excel_report(metrics_list, time_series_list, filename):
    """Create a professionally formatted, presentation-grade Excel spreadsheet."""
    print(f"\n[Excel] Writing results to Excel file: {filename}")
    wb = openpyxl.Workbook()
    
    # Define styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_bold_data = Font(name="Segoe UI", size=10, bold=True)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate Gray
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")   # Very light blue-gray
    fill_summary = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Muted gray highlight
    
    border_thin = Side(border_style="thin", color="CBD5E1")
    border_thick = Side(border_style="medium", color="475569")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_heavy_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thick)
    
    # ----------------------------------------------------
    # TAB 1: SUMMARY DASHBOARD
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1["A1"] = "HealthSense AI API Load Testing Report"
    ws1["A1"].font = font_title
    ws1["A2"] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: {API_URL}"
    ws1["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    
    # Meta Configuration Card
    ws1["A4"] = "Test Configuration"
    ws1["A4"].font = font_section
    
    ws1["A5"] = "Parameter"
    ws1["B5"] = "Value"
    for col in ["A", "B"]:
        ws1[f"{col}5"].font = font_header
        ws1[f"{col}5"].fill = fill_header
        ws1[f"{col}5"].alignment = align_center
        ws1[f"{col}5"].border = cell_border
        
    config_params = [
        ("Concurrency (VUs)", CONCURRENCY),
        ("Duration Per Test", f"{DURATION_SECONDS} seconds"),
        ("Lightweight Endpoint", "GET / (Framework Baseline)"),
        ("Database Endpoint", "GET /user/profile (Auth & DB Baseline)"),
        ("Test Account Used", TEST_EMAIL),
    ]
    
    for idx, (param, val) in enumerate(config_params):
        row = 6 + idx
        ws1[f"A{row}"] = param
        ws1[f"B{row}"] = val
        ws1[f"A{row}"].font = font_data
        ws1[f"A{row}"].border = cell_border
        ws1[f"B{row}"].font = font_bold_data
        ws1[f"B{row}"].border = cell_border
        ws1[f"B{row}"].alignment = align_right
        
    # Comparative Results Table
    ws1["A13"] = "Performance Summary Comparison"
    ws1["A13"].font = font_section
    
    headers = [
        "Scenario Name", "Total Requests", "Success Rate", "Avg RPS",
        "Avg Latency (ms)", "Min Latency (ms)", "Max Latency (ms)",
        "p50 (Median)", "p95 Latency", "p99 Latency"
    ]
    
    for col_idx, h in enumerate(headers):
        col_letter = get_column_letter(col_idx + 1)
        ws1[f"{col_letter}14"] = h
        ws1[f"{col_letter}14"].font = font_header
        ws1[f"{col_letter}14"].fill = fill_header
        ws1[f"{col_letter}14"].alignment = align_center
        ws1[f"{col_letter}14"].border = cell_border
        
    for r_idx, metrics in enumerate(metrics_list):
        row = 15 + r_idx
        ws1.cell(row=row, column=1, value=metrics["name"]).font = font_bold_data
        ws1.cell(row=row, column=2, value=metrics["total_requests"]).number_format = '#,##0'
        ws1.cell(row=row, column=3, value=metrics["success_rate"] / 100.0).number_format = '0.0%'
        ws1.cell(row=row, column=4, value=metrics["rps"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=5, value=metrics["avg_latency"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=6, value=metrics["min_latency"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=7, value=metrics["max_latency"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=8, value=metrics["p50"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=9, value=metrics["p95"]).number_format = '#,##0.0'
        ws1.cell(row=row, column=10, value=metrics["p99"]).number_format = '#,##0.0'
        
        for c_idx in range(1, 11):
            cell = ws1.cell(row=row, column=c_idx)
            cell.font = font_bold_data if c_idx == 1 else font_data
            cell.border = cell_border
            if c_idx > 1:
                cell.alignment = align_right
            else:
                cell.alignment = align_left
                
            # Zebra stripe the rows
            if r_idx % 2 == 1:
                cell.fill = fill_zebra
                
    # Auto-fit columns for Summary tab
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ----------------------------------------------------
    # TAB 2: DETAILED TIME-SERIES DATA & CHARTS
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Time-Series Data")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "Scenario A: Public Root (/)"
    ws2["A1"].font = font_section
    ws2["E1"] = "Scenario B: Authenticated Profile (/user/profile)"
    ws2["E1"].font = font_section
    
    headers_ts = ["Second", "Requests (RPS)", "Avg Latency (ms)"]
    
    # Headers for Scenario A
    for c_idx, h in enumerate(headers_ts):
        col_letter = get_column_letter(c_idx + 1)
        ws2[f"{col_letter}2"] = h
        ws2[f"{col_letter}2"].font = font_header
        ws2[f"{col_letter}2"].fill = fill_header
        ws2[f"{col_letter}2"].alignment = align_center
        ws2[f"{col_letter}2"].border = cell_border
        
    # Headers for Scenario B
    for c_idx, h in enumerate(headers_ts):
        col_letter = get_column_letter(c_idx + 5)
        ws2[f"{col_letter}2"] = h
        ws2[f"{col_letter}2"].font = font_header
        ws2[f"{col_letter}2"].fill = fill_header
        ws2[f"{col_letter}2"].alignment = align_center
        ws2[f"{col_letter}2"].border = cell_border
        
    # Populate Time-Series data
    ts_a = time_series_list[0]
    ts_b = time_series_list[1]
    
    for i in range(DURATION_SECONDS):
        row = 3 + i
        
        # Scenario A
        sec_a, reqs_a, lat_a = ts_a[i]
        ws2.cell(row=row, column=1, value=sec_a).alignment = align_center
        ws2.cell(row=row, column=2, value=reqs_a).number_format = '#,##0'
        ws2.cell(row=row, column=3, value=lat_a).number_format = '#,##0.0'
        
        # Scenario B
        sec_b, reqs_b, lat_b = ts_b[i]
        ws2.cell(row=row, column=5, value=sec_b).alignment = align_center
        ws2.cell(row=row, column=6, value=reqs_b).number_format = '#,##0'
        ws2.cell(row=row, column=7, value=lat_b).number_format = '#,##0.0'
        
        for c_idx in [1, 2, 3, 5, 6, 7]:
            cell = ws2.cell(row=row, column=c_idx)
            cell.font = font_data
            cell.border = cell_border
            if c_idx in [2, 3, 6, 7]:
                cell.alignment = align_right
            if i % 2 == 1:
                cell.fill = fill_zebra
                
    # Auto-fit columns for Time-Series tab
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        col = ws2[col_letter]
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Create Line Charts for visualization!
    # Chart 1: Requests Per Second (RPS) comparison
    chart_rps = LineChart()
    chart_rps.title = "Throughput over Time (RPS)"
    chart_rps.style = 13
    chart_rps.y_axis.title = "Requests / Second"
    chart_rps.x_axis.title = "Elapsed Time (Seconds)"
    chart_rps.width = 16
    chart_rps.height = 10
    
    data_rps_a = Reference(ws2, min_col=2, min_row=2, max_row=2+DURATION_SECONDS)
    data_rps_b = Reference(ws2, min_col=6, min_row=2, max_row=2+DURATION_SECONDS)
    cats = Reference(ws2, min_col=1, min_row=3, max_row=2+DURATION_SECONDS)
    
    chart_rps.add_data(data_rps_a, titles_from_data=True)
    chart_rps.add_data(data_rps_b, titles_from_data=True)
    chart_rps.set_categories(cats)
    
    # Chart 2: Average Latency comparison
    chart_lat = LineChart()
    chart_lat.title = "Average Response Time over Time (Latency)"
    chart_lat.style = 13
    chart_lat.y_axis.title = "Latency (Milliseconds)"
    chart_lat.x_axis.title = "Elapsed Time (Seconds)"
    chart_lat.width = 16
    chart_lat.height = 10
    
    data_lat_a = Reference(ws2, min_col=3, min_row=2, max_row=2+DURATION_SECONDS)
    data_lat_b = Reference(ws2, min_col=7, min_row=2, max_row=2+DURATION_SECONDS)
    
    chart_lat.add_data(data_lat_a, titles_from_data=True)
    chart_lat.add_data(data_lat_b, titles_from_data=True)
    chart_lat.set_categories(cats)
    
    # Place charts in Summary Dashboard
    ws1.add_chart(chart_rps, "A19")
    ws1.add_chart(chart_lat, "E19")
    
    # Save Workbook
    wb.save(filename)
    print(f"[Excel] Excel report generated successfully!")

async def main():
    print("Initializing Baseline/Load Testing...")
    
    # Start the local FastAPI server
    server_process, log_file = start_backend_server()
    
    try:
        # Wait for server to boot up
        server_ready = await wait_for_server(timeout=15.0)
        if not server_ready:
            print("[CRITICAL] Server failed to start or become responsive. Check uvicorn logs.")
            return
            
        # Get DB status from startup log if possible
        # Give it a second to connect to the db
        await asyncio.sleep(2.0)
        
        # Authenticate
        token = await authenticate_user()
        if not token:
            print("[CRITICAL] Authentication failed. Cannot run Scenario B. Aborting.")
            return
            
        headers_auth = {"Authorization": f"Bearer {token}"}
        headers_pub = {}
        
        # Scenario A: Public Root Endpoint (GET /)
        results_a = await run_scenario(
            "Scenario A (Public Framework Overhead)", 
            "/", 
            headers_pub, 
            DURATION_SECONDS, 
            CONCURRENCY
        )
        
        # Give the server 3 seconds to recover/cooldown between runs
        print("\n[Cooldown] Pausing for 3 seconds...")
        await asyncio.sleep(3.0)
        
        # Scenario B: Authenticated Profile Endpoint (GET /user/profile)
        results_b = await run_scenario(
            "Scenario B (Auth & Database Profile Query)", 
            "/user/profile", 
            headers_auth, 
            DURATION_SECONDS, 
            CONCURRENCY
        )
        
        # Stop background server safely
        print("\n[Server] Terminating backend FastAPI server...")
        server_process.terminate()
        server_process.wait()
        log_file.close()
        print("[Server] Server process terminated successfully.")
        
        # Analyze metrics
        metrics_a = compute_metrics("Scenario A: Public Root (/) - Framework Overhead", results_a, DURATION_SECONDS)
        metrics_b = compute_metrics("Scenario B: Authenticated (/user/profile) - DB & JWT", results_b, DURATION_SECONDS)
        
        print("\n" + "="*50)
        print(" RESULTS OVERVIEW")
        print("="*50)
        for m in [metrics_a, metrics_b]:
            print(f"Scenario: {m['name']}")
            print(f"  Total Requests: {m['total_requests']}")
            print(f"  Success Rate:   {m['success_rate']:.2f}%")
            print(f"  Average RPS:    {m['rps']:.2f} req/sec")
            print(f"  Response Latencies:")
            print(f"    Min:   {m['min_latency']:.1f}ms")
            print(f"    Avg:   {m['avg_latency']:.1f}ms")
            print(f"    Max:   {m['max_latency']:.1f}ms")
            print(f"    p50:   {m['p50']:.1f}ms")
            print(f"    p95:   {m['p95']:.1f}ms")
            print(f"    p99:   {m['p99']:.1f}ms")
            print("-"*30)
            
        # Get Time-Series Data
        ts_a = get_time_series_data(results_a, DURATION_SECONDS)
        ts_b = get_time_series_data(results_b, DURATION_SECONDS)
        
        # Save Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = os.path.join(PROJECT_ROOT, "load_testing", f"Load_Test_Report_HealthSense_{timestamp}.xlsx")
        generate_excel_report([metrics_a, metrics_b], [ts_a, ts_b], report_filename)
        
        # Write a copy to a standard name for verification
        stable_filename = os.path.join(PROJECT_ROOT, "load_testing", "Load_Test_Report_Latest.xlsx")
        import shutil
        shutil.copyfile(report_filename, stable_filename)
        print(f"[Excel] Copied latest report to stable path: {stable_filename}")
        
    except Exception as e:
        print(f"[CRITICAL] Error running load test orchestration: {e}")
        import traceback
        traceback.print_exc()
        if server_process.poll() is None:
            server_process.terminate()
            server_process.wait()
            log_file.close()

if __name__ == "__main__":
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(main())
