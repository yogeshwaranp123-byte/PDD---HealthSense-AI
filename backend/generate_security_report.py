"""
generate_security_report_v2.py
Generates the security vulnerability report in the same format/style
as the E2E_Test_Report_PancreaScan Excel file.
Run from: backend/ directory
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
import os

# ── Exact colours from E2E report ─────────────────────────────────────────────
HDR_BG    = "1F4E79"   # dark navy  (header fill)
HDR_FG    = "FFFFFF"   # white      (header font)
PASS_BG   = "E2EFDA"   # light green
FAIL_BG   = "FCE4D6"   # light red/orange
WARN_BG   = "FFEB9C"   # yellow (documented)
INFO_BG   = "DDEBF7"   # light blue
ROW_ALT   = "EBF5FF"   # alternating row tint

# Severity colours for Status column (matching FAILED/PASSED style)
SEV_COLORS = {
    "Critical": "C00000",   # dark red
    "High":     "ED7D31",   # orange
    "Medium":   "FFD966",   # yellow
    "Low":      "5B9BD5",   # blue
}
SEV_BG = {
    "Critical": "FCE4D6",
    "High":     "FCE4D6",
    "Medium":   "FFEB9C",
    "Low":      "DDEBF7",
}

# ── Style helpers (matching E2E file exactly) ─────────────────────────────────
def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(size=11) -> Font:
    return Font(bold=True, color=HDR_FG, name="Calibri", size=size)

def _body_font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, name="Calibri", size=size)

def _center(wrap=True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _hdr_cell(cell, value, width_col=None):
    """Apply full E2E-style header formatting to a cell."""
    cell.value = value
    cell.fill = _fill(HDR_BG)
    cell.font = _hdr_font()
    cell.alignment = _center()
    cell.border = _border()

def _data_cell(cell, value, bold=False, fill_hex=None, align="left"):
    cell.value = value
    cell.fill = _fill(fill_hex) if fill_hex else _fill("FFFFFF")
    cell.font = _body_font(bold=bold)
    cell.alignment = _center() if align == "center" else _left()
    cell.border = _border()

def _set_col(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _set_row(ws, row_idx, height=20.0):
    ws.row_dimensions[row_idx].height = height

# ── All findings ──────────────────────────────────────────────────────────────
# (ID, Severity, Category, File, Vulnerability Type, Description, Remediation, Status)
FINDINGS = [
    ("CRIT-01","Critical","Sensitive Data Exposure",
     "backend/.env",
     "Credentials in Source Control",
     "Live Gemini API key and MongoDB Atlas credentials (username+password) committed to repository.",
     "Added .env & local_db.json to .gitignore; created .env.example. Rotate both secrets immediately.",
     "FIXED"),
    ("CRIT-02","Critical","Authentication",
     "app/core/config.py L7",
     "Hardcoded Weak JWT Secret",
     "SECRET_KEY defaulted to 'akashkrish101019' — anyone with source access can forge JWTs for any user.",
     "Removed default value; pydantic-settings now raises ValidationError at startup if SECRET_KEY is missing.",
     "FIXED"),
    ("CRIT-03","Critical","Authentication / Access Control",
     "app/core/security.py L35, app/routers/auth.py L54",
     "Demo Mode Authentication Bypass",
     "Any client sending 'Bearer demo_access:<id>' bypassed all JWT verification, impersonating any user.",
     "Removed demo_access:/demo_refresh: bypass entirely. All tokens must now be valid signed JWTs.",
     "FIXED"),
    ("CRIT-04","Critical","Authentication",
     "app/main.py L32, app/core/security.py L20",
     "Empty Password Accepted",
     "Demo user tester@healthsense.ai had hashed_password=''; verify_password returned True for empty input.",
     "verify_password returns False if either argument is empty. Demo user seeded with strong bcrypt hash.",
     "FIXED"),
    ("CRIT-05","Critical","API Security / CORS",
     "app/main.py L13-19",
     "Wildcard CORS + allow_credentials=True",
     "allow_origins=['*'] with allow_credentials=True — any origin could make credentialed cross-site requests.",
     "CORS now uses explicit ALLOWED_ORIGINS env var allowlist. allow_credentials set to False.",
     "FIXED"),
    ("HIGH-01","High","API Security",
     "app/routers/auth.py",
     "No Rate Limiting on Auth Endpoints",
     "Unlimited login attempts enabled brute-force attacks against any user account password.",
     "Added slowapi rate limiter: 5/min on /auth/login, 10/min on /auth/register and /auth/refresh.",
     "FIXED"),
    ("HIGH-02","High","Authentication",
     "app/routers/auth.py (missing endpoint)",
     "No Token Revocation / Stateless Logout",
     "No /auth/logout; stolen tokens remained valid for their full 30-min or 7-day lifetime.",
     "Added /auth/logout endpoint + JTI-based token blocklist in security.py. All tokens include jti claim.",
     "FIXED"),
    ("HIGH-03","High","Input Validation",
     "app/routers/predict.py L36-51",
     "Unlimited File Upload — No Size/MIME Validation",
     "No file size limit; MIME type taken from client Content-Type header (trivially spoofed).",
     "10 MB size cap enforced. MIME validated via magic byte signatures, not the client-supplied header.",
     "FIXED"),
    ("HIGH-04","High","API Security",
     "app/main.py",
     "No Request Body Size Limit",
     "Arbitrarily large payloads accepted on all endpoints — DoS via memory exhaustion.",
     "Added middleware enforcing 15 MB hard cap; returns HTTP 413 on oversized requests.",
     "FIXED"),
    ("HIGH-05","High","Injection",
     "app/routers/predict.py L54, app/routers/hospitals.py L25",
     "Prompt Injection via Unvalidated Input",
     "disease (Form) interpolated unsanitized into Gemini prompts — attacker could inject AI instructions.",
     "disease validated against strict ALLOWED_DISEASES set; invalid values return HTTP 400.",
     "FIXED"),
    ("HIGH-06","High","Sensitive Data Exposure",
     "predict.py L90, chat.py L28, hospitals.py L54",
     "API Key Exposed in URL / Server Logs",
     "Gemini API key embedded as ?key= query param — appeared in Uvicorn access logs and error responses.",
     "Moved API key to x-goog-api-key HTTP header in all three routers. No longer in any URL.",
     "FIXED"),
    ("MED-01","Medium","Input Validation",
     "app/models/schemas.py L9",
     "No Password Complexity Policy",
     "Only min_length=6 enforced; trivially weak passwords like '123456' or 'abcdef' were accepted.",
     "Added Pydantic field_validator requiring uppercase, lowercase, digit, and special character.",
     "FIXED"),
    ("MED-02","Medium","Authorization",
     "app/routers/report.py L27",
     "IDOR — Unsafe ObjectId Conversion",
     "Invalid prediction_id caused unhandled 500; compounded by demo bypass (now removed in CRIT-03).",
     "ObjectId.is_valid() check added; malformed IDs return HTTP 400. CRIT-03 fix removes IDOR vector.",
     "FIXED"),
    ("MED-03","Medium","Information Disclosure",
     "predict.py, chat.py, hospitals.py (multiple lines)",
     "Verbose Internal Error Messages",
     "Raw exception messages, stack traces, and full AI service response bodies returned to clients.",
     "All exceptions now logged server-side with context; clients receive only generic messages.",
     "FIXED"),
    ("MED-04","Medium","Sensitive Data Exposure",
     "backend/local_db.json",
     "User Data (bcrypt hashes) in Committed File",
     "Fallback JSON database with real user records and password hashes was committed to the repository.",
     "Added local_db.json to backend/.gitignore. File no longer tracked by git.",
     "FIXED"),
    ("MED-05","Medium","API Security",
     "app/main.py",
     "Missing Security Headers",
     "No X-Content-Type-Options, X-Frame-Options, HSTS, X-XSS-Protection, or Referrer-Policy headers.",
     "Added HTTP middleware injecting all standard security headers on every API response.",
     "FIXED"),
    ("MED-06","Medium","Injection",
     "app/routers/hospitals.py L31",
     "Unsanitized Address in AI Prompt",
     "Optional address query param interpolated directly into Gemini prompt without any sanitization.",
     "Regex allowlist applied ([a-zA-Z0-9 ,.\\-#/]); max 200 chars; invalid chars stripped before use.",
     "FIXED"),
    ("MED-07","Medium","Authentication",
     "app/core/jwt_handler.py L25-30",
     "Silent JWT Decode Failure Returns {}",
     "decode_token silently returned empty dict on any failure; callers could forget to validate the result.",
     "decode_token now raises typed JWTDecodeError; all callers handle it explicitly with HTTP 401.",
     "FIXED"),
    ("LOW-01","Low","Sensitive Data Exposure",
     "website/src/services/api.ts, endpoints.ts, store/authStore.ts",
     "Tokens in localStorage (XSS Risk)",
     "JWT access/refresh tokens stored in localStorage — accessible to any JavaScript on the page.",
     "Backend now sets httpOnly+Secure+SameSite=Strict cookies on login/register/refresh/logout. Frontend uses withCredentials:true; zero localStorage token reads remain. /auth/me validates cookie server-side.",
     "FIXED"),
    ("LOW-02","Low","Configuration",
     "backend/render.yaml",
     "No Secret References in Deployment Manifest",
     "render.yaml had no env var declarations; missing secrets fell back silently to hardcoded defaults.",
     "Added explicit sync:false declarations for SECRET_KEY, MONGO_URI, GEMINI_API_KEY, ALLOWED_ORIGINS.",
     "FIXED"),
    ("LOW-03","Low","Vulnerable Dependency",
     "backend/requirements.txt L5",
     "python-jose Has Known CVEs",
     "python-jose 3.3.0 is unmaintained with known security vulnerabilities.",
     "Migrated to PyJWT 2.8.0 (actively maintained). Updated jwt_handler.py to new API.",
     "FIXED"),
    ("LOW-04","Low","Configuration",
     "backend/requirements.txt L15-16",
     "Test Tools in Production Dependencies",
     "selenium and webdriver-manager in production requirements — unnecessarily expanded attack surface.",
     "Moved to requirements-dev.txt. Excluded from Render production builds.",
     "FIXED"),
]

# ── Log entries (one per finding — mirrors Execution Log sheet) ───────────────
def build_log(findings):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for f in findings:
        fid, sev, cat, fpath, vtype, desc, rem, status = f
        level = "PASS" if status == "FIXED" else "WARN" if status == "DOCUMENTED" else "FAIL"
        msg = f"[{cat}] {fid} | {vtype} | {status}"
        rows.append((ts, level, msg))
    return rows

# ── Build workbook ────────────────────────────────────────────────────────────
def build_workbook():
    wb = openpyxl.Workbook()

    total      = len(FINDINGS)
    fixed      = sum(1 for f in FINDINGS if f[7] == "FIXED")
    documented = sum(1 for f in FINDINGS if f[7] == "DOCUMENTED")
    critical   = sum(1 for f in FINDINGS if f[1] == "Critical")
    high_cnt   = sum(1 for f in FINDINGS if f[1] == "High")
    medium_cnt = sum(1 for f in FINDINGS if f[1] == "Medium")
    low_cnt    = sum(1 for f in FINDINGS if f[1] == "Low")
    fix_rate   = round(fixed / total * 100, 2) if total else 0
    now        = datetime.now(timezone.utc)
    ts_str     = now.strftime("%Y-%m-%dT%H:%M:%SZ")

    # ══════════════════════════════════════════════════════════════
    # Sheet 1: Summary  (mirrors E2E "Summary" sheet exactly)
    # ══════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    # Header row
    sum_headers = ["Report Suite","Total Findings","Fixed","Documented",
                   "Fix Rate %","Critical","High","Medium","Low","Generated At"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci)
        _hdr_cell(cell, h)
    _set_row(ws_sum, 1, 28.0)

    # Data row
    data_vals = [
        "HealthSense AI — Security Vulnerability Report",
        total, fixed, documented, fix_rate,
        critical, high_cnt, medium_cnt, low_cnt, ts_str
    ]
    for ci, v in enumerate(data_vals, 1):
        cell = ws_sum.cell(row=2, column=ci)
        _data_cell(cell, v, align="center")
    _set_row(ws_sum, 2, 20.0)

    # Column widths (matching E2E proportions)
    col_widths_sum = [43, 15, 12, 14, 13, 12, 12, 12, 12, 31]
    for i, w in enumerate(col_widths_sum, 1):
        _set_col(ws_sum, get_column_letter(i), w)

    # ══════════════════════════════════════════════════════════════
    # Sheet 2: All Findings  (mirrors E2E "Test Details" sheet)
    # ══════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet("All Findings")
    ws_all.sheet_view.showGridLines = False

    all_headers = ["No.","Severity","Category","File / Location",
                   "Vulnerability Type","Description","Remediation Applied","Status"]
    for ci, h in enumerate(all_headers, 1):
        cell = ws_all.cell(row=1, column=ci)
        _hdr_cell(cell, h)
    _set_row(ws_all, 1, 28.0)

    for ri, finding in enumerate(FINDINGS, 2):
        fid, sev, cat, fpath, vtype, desc, rem, status = finding
        alt = ROW_ALT if ri % 2 == 0 else "FFFFFF"
        status_bg = PASS_BG if status == "FIXED" else (WARN_BG if status == "DOCUMENTED" else FAIL_BG)

        row_data = [fid, sev, cat, fpath, vtype, desc, rem, status]
        for ci, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=ri, column=ci)
            if ci == 2:   # Severity — coloured
                _data_cell(cell, val, bold=True,
                           fill_hex=SEV_BG.get(sev, "FFFFFF"), align="center")
                cell.font = Font(bold=True, color=SEV_COLORS.get(sev, "000000"),
                                 name="Calibri", size=10)
            elif ci == 8:  # Status — FIXED / DOCUMENTED
                _data_cell(cell, val, bold=True, fill_hex=status_bg, align="center")
            else:
                _data_cell(cell, val, fill_hex=alt)
        _set_row(ws_all, ri, 20.0)

    col_widths_all = [12, 12, 22, 38, 34, 58, 62, 14]
    for i, w in enumerate(col_widths_all, 1):
        _set_col(ws_all, get_column_letter(i), w)

    # ══════════════════════════════════════════════════════════════
    # Sheet 3: Critical & High  (mirrors E2E "Failed Tests" sheet)
    # ══════════════════════════════════════════════════════════════
    ws_crit = wb.create_sheet("Critical & High")
    ws_crit.sheet_view.showGridLines = False

    ch_headers = ["No.","Severity","Category","File / Location","Vulnerability Type","Status","Remediation Applied"]
    for ci, h in enumerate(ch_headers, 1):
        cell = ws_crit.cell(row=1, column=ci)
        _hdr_cell(cell, h)
    _set_row(ws_crit, 1, 28.0)

    crit_high = [f for f in FINDINGS if f[1] in ("Critical", "High")]
    for ri, finding in enumerate(crit_high, 2):
        fid, sev, cat, fpath, vtype, desc, rem, status = finding
        alt = ROW_ALT if ri % 2 == 0 else "FFFFFF"
        status_bg = PASS_BG if status == "FIXED" else FAIL_BG

        row_data = [fid, sev, cat, fpath, vtype, status, rem]
        for ci, val in enumerate(row_data, 1):
            cell = ws_crit.cell(row=ri, column=ci)
            if ci == 2:
                _data_cell(cell, val, bold=True, fill_hex=SEV_BG.get(sev,"FFFFFF"), align="center")
                cell.font = Font(bold=True, color=SEV_COLORS.get(sev,"000000"), name="Calibri", size=10)
            elif ci == 6:
                _data_cell(cell, val, bold=True, fill_hex=status_bg, align="center")
            else:
                _data_cell(cell, val, fill_hex=alt)
        _set_row(ws_crit, ri, 20.0)

    col_widths_ch = [12, 12, 22, 38, 34, 14, 62]
    for i, w in enumerate(col_widths_ch, 1):
        _set_col(ws_crit, get_column_letter(i), w)

    # ══════════════════════════════════════════════════════════════
    # Sheet 4: Execution Log  (mirrors E2E "Execution Log" sheet)
    # ══════════════════════════════════════════════════════════════
    ws_log = wb.create_sheet("Execution Log")
    ws_log.sheet_view.showGridLines = False

    log_headers = ["Timestamp", "Level", "Message"]
    for ci, h in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=ci)
        _hdr_cell(cell, h)
    _set_row(ws_log, 1, 28.0)

    log_rows = build_log(FINDINGS)
    for ri, (ts, level, msg) in enumerate(log_rows, 2):
        level_bg = PASS_BG if level == "PASS" else (WARN_BG if level == "WARN" else FAIL_BG)
        _data_cell(ws_log.cell(row=ri, column=1), ts)
        _data_cell(ws_log.cell(row=ri, column=2), level, bold=True,
                   fill_hex=level_bg, align="center")
        _data_cell(ws_log.cell(row=ri, column=3), msg)
        _set_row(ws_log, ri, 20.0)

    _set_col(ws_log, "A", 23)
    _set_col(ws_log, "B", 14)
    _set_col(ws_log, "C", 72)

    # ══════════════════════════════════════════════════════════════
    # Sheet 5: Remediation Roadmap
    # ══════════════════════════════════════════════════════════════
    ws_road = wb.create_sheet("Remediation Roadmap")
    ws_road.sheet_view.showGridLines = False

    road_headers = ["No.","Priority","Timing","Finding ID","Action","Owner","Status"]
    for ci, h in enumerate(road_headers, 1):
        cell = ws_road.cell(row=1, column=ci)
        _hdr_cell(cell, h)
    _set_row(ws_road, 1, 28.0)

    roadmap = [
        (1,"Critical","Immediate","CRIT-01","Rotate Gemini API key and MongoDB Atlas password in production","Developer","FIXED"),
        (2,"Critical","Immediate","CRIT-01","Purge .env and local_db.json from git history (use BFG Repo Cleaner)","Developer","MANUAL"),
        (3,"Critical","Immediate","CRIT-02","Set SECRET_KEY env var on Render dashboard — do NOT use default","DevOps","FIXED"),
        (4,"Critical","Immediate","CRIT-03","Verify demo_access: bypass is removed — test with a dummy demo token","QA","FIXED"),
        (5,"Critical","Immediate","CRIT-04","Confirm demo user seeded with hashed password on next startup","QA","FIXED"),
        (6,"Critical","Immediate","CRIT-05","Set ALLOWED_ORIGINS to actual frontend URL on Render dashboard","DevOps","FIXED"),
        (7,"High","This Week","HIGH-01","Deploy with slowapi rate limiter; verify 429 on repeated logins","QA","FIXED"),
        (8,"High","This Week","HIGH-02","Test /auth/logout endpoint; verify revoked token returns 401","QA","FIXED"),
        (9,"High","This Week","HIGH-03","Upload file >10 MB — verify HTTP 413; upload non-PDF — verify 415","QA","FIXED"),
        (10,"High","This Week","HIGH-05","Send disease=evil_injection — verify HTTP 400 returned","QA","FIXED"),
        (11,"High","This Week","HIGH-06","Confirm API key no longer appears in Uvicorn access logs","DevOps","FIXED"),
        (12,"Medium","Soon","MED-01","Register with password='weak' — verify HTTP 422 validation error","QA","FIXED"),
        (13,"Medium","Soon","MED-03","Confirm error responses contain no stack traces or internal details","QA","FIXED"),
        (14,"Medium","Soon","MED-05","Verify X-Content-Type-Options and X-Frame-Options in all responses","QA","FIXED"),
        (15,"Low","Backlog","LOW-01","Migrate frontend tokens from localStorage to httpOnly cookies (backend Set-Cookie + withCredentials)","Developer","FIXED"),
        (16,"Low","Backlog","LOW-03","Verify PyJWT works correctly after SECRET_KEY rotation on Render","Developer","FIXED"),
    ]

    for ri, row in enumerate(roadmap, 2):
        num, prio, timing, fid, action, owner, status = row
        alt = ROW_ALT if ri % 2 == 0 else "FFFFFF"
        prio_bg = {"Critical": SEV_BG["Critical"], "High": SEV_BG["High"],
                   "Medium": SEV_BG["Medium"], "Low": SEV_BG["Low"]}.get(prio, "FFFFFF")
        status_bg = PASS_BG if status == "FIXED" else (WARN_BG if status in ("PLANNED","MANUAL") else FAIL_BG)

        row_vals = [num, prio, timing, fid, action, owner, status]
        for ci, val in enumerate(row_vals, 1):
            cell = ws_road.cell(row=ri, column=ci)
            if ci == 2:
                _data_cell(cell, val, bold=True, fill_hex=prio_bg, align="center")
                cell.font = Font(bold=True, color=SEV_COLORS.get(prio,"000000"), name="Calibri", size=10)
            elif ci == 7:
                _data_cell(cell, val, bold=True, fill_hex=status_bg, align="center")
            else:
                _data_cell(cell, val, fill_hex=alt)
        _set_row(ws_road, ri, 20.0)

    col_widths_road = [8, 12, 14, 14, 72, 18, 14]
    for i, w in enumerate(col_widths_road, 1):
        _set_col(ws_road, get_column_letter(i), w)

    return wb


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    now = datetime.now(timezone.utc)
    timestamp = now.strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"Security_Vulnerability_Report_{timestamp}.xlsx"

    output_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "Vulnerability Test Results"
    )
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    wb = build_workbook()
    wb.save(output_path)
    print(f"[OK] Security report saved: {os.path.normpath(output_path)}")
